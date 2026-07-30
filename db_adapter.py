"""
Database abstraction layer per DVR Checklist Analyzer.
Supporta due backend:
- 'local': Excel locale (comportamento attuale)
- 'supabase': Supabase PostgreSQL + Storage
"""
from abc import ABC, abstractmethod
from typing import List, Dict, Optional, Any
import re
import os
import json


class DatabaseAdapter(ABC):
    @abstractmethod
    def get_checklist_items(self) -> List[Dict[str, Any]]:
        pass

    @abstractmethod
    def get_all_dvr_risks(self) -> List[Dict[str, Any]]:
        pass

    @abstractmethod
    def match_risks(self, checked_keys: List[str]) -> List[Dict[str, Any]]:
        pass

    @abstractmethod
    def add_database_risk(self, risk: Dict[str, Any]) -> Dict[str, Any]:
        pass


class ExcelAdapter(DatabaseAdapter):
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._checklist_cache = None
        self._risks_cache = None
        self._images_cache = None
        self._users_path = os.path.join(os.path.expanduser("~"), "Documents", "CONTEA DVR Analyzer", "users.json")

    def _get_images_map(self):
        if self._images_cache is not None:
            return self._images_cache
        import zipfile
        import xml.etree.ElementTree as ET
        import base64

        self._images_cache = {}
        if not __import__('os').path.exists(self.db_path):
            return self._images_cache

        try:
            import zipfile
            import xml.etree.ElementTree as ET
            import base64

            with zipfile.ZipFile(self.db_path, 'r') as z:
                if 'xl/drawings/drawing9.xml' not in z.namelist():
                    return self._images_cache
                rels = z.read('xl/drawings/_rels/drawing9.xml.rels').decode('utf-8')
                root = ET.fromstring(rels)
                rid_to_img = {}
                for rel in root:
                    rid_to_img[rel.attrib['Id']] = rel.attrib['Target'].replace('../', 'xl/')

                d9 = z.read('xl/drawings/drawing9.xml').decode('utf-8')
                root_d9 = ET.fromstring(d9)
                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                }
                for anchor in root_d9.findall('xdr:oneCellAnchor', ns):
                    frm = anchor.find('xdr:from', ns)
                    blip = anchor.find('.//a:blip', ns)
                    if frm is not None and blip is not None:
                        r = frm.find('xdr:row', ns)
                        embed = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        if r is not None and embed:
                            row_num = int(r.text) + 1
                            p = rid_to_img.get(embed)
                            if p and p in z.namelist():
                                b = z.read(p)
                                mime = 'image/png' if p.endswith('.png') else 'image/jpeg'
                                b64 = base64.b64encode(b).decode('utf-8')
                                self._images_cache[row_num] = f"data:{mime};base64,{b64}"
        except Exception as e:
            print("Errore estrazione immagini Excel:", e)
        return self._images_cache

    def _read_checklist(self):
        if self._checklist_cache is not None:
            return self._checklist_cache
        self._checklist_cache = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.db_path, data_only=True)
            if 'CHECK LIST' not in wb.sheetnames:
                return self._checklist_cache
            ws = wb['CHECK LIST']
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx < 4:
                    continue
                cat = row[0] or ""
                item = row[1]
                if item is None:
                    continue
                is_cb = isinstance(row[2], bool)
                key = row[6] if row[6] is not None else f"{cat} {item}"
                self._checklist_cache.append({
                    'category': str(cat).strip(),
                    'item': str(item).strip(),
                    'key': str(key).strip(),
                    'is_checkbox': is_cb
                })
        except Exception as e:
            print("Errore lettura checklist:", e)
        return self._checklist_cache

    def _read_risks(self):
        if self._risks_cache is not None:
            return self._risks_cache
        self._risks_cache = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.db_path, data_only=True)
            if 'DVR - RISCHI' not in wb.sheetnames:
                return self._risks_cache
            ws = wb['DVR - RISCHI']
            imgs = self._get_images_map()
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx < 2:
                    continue
                rn = r_idx + 1
                luogo = row[0] or ""
                rischio = row[1] or ""
                azione = row[2] or ""
                entro = row[4] or ""
                key = row[5] or ""
                raw_ord = row[7] if len(row) > 7 else None
                try:
                    ordine = int(raw_ord) if raw_ord is not None else 999999
                except:
                    ordine = 999999
                img = imgs.get(rn)
                if luogo or rischio or azione:
                    self._risks_cache.append({
                        'luogo': str(luogo).strip(), 'rischio': str(rischio).strip(),
                        'azione': str(azione).strip(), 'entro': str(entro).strip(),
                        'key': str(key).strip(), 'ordine': ordine, 'image': img
                    })
        except Exception as e:
            print("Errore lettura rischi:", e)
        return self._risks_cache

    def get_checklist_items(self):
        return self._read_checklist()

    def get_all_dvr_risks(self):
        risks = self._read_risks()
        risks.sort(key=lambda x: x.get('ordine', 999999))
        return risks

    def match_risks(self, checked_keys: List[str]) -> List[Dict[str, Any]]:
        risks = self.get_all_dvr_risks()
        matched = []
        used = set()
        norm_ck = [self._normalize(k) for k in checked_keys]
        for i, ck in enumerate(checked_keys):
            n = norm_ck[i]
            if not n:
                continue
            for j, r in enumerate(risks):
                if j in used:
                    continue
                nr = self._normalize(r['key'])
                if nr == n or n in nr or nr in n:
                    matched.append({**r, 'key': ck})
                    used.add(j)
                    break
        matched.sort(key=lambda x: x.get('ordine', 999999))
        return matched

    def add_database_risk(self, risk: Dict[str, Any]) -> Dict[str, Any]:
        import openpyxl
        wb = openpyxl.load_workbook(self.db_path)
        ws = wb['DVR - RISCHI']
        ord_val = risk.get('ordine', 999999)
        try:
            ordine = int(ord_val) if ord_val is not None else 999999
        except:
            ordine = 999999
        key_val = risk.get('rischio') or risk.get('luogo') or ""
        new_row = [
            risk.get('luogo', ''),
            risk.get('rischio', ''),
            risk.get('azione', ''),
            None,
            risk.get('entro', ''),
            key_val,
            'FALSO',
            ordine
        ]
        ws.append(new_row)
        wb.save(self.db_path)
        self._risks_cache = None
        return {"success": True, "message": "Criticità salvata in Excel."}

    @staticmethod
    def _normalize(k):
        if not k:
            return ""
        k = str(k).lower().strip()
        k = re.sub(r'\(.*?\)', '', k)
        k = re.sub(r'[^a-z0-9\s]', ' ', k)
        return " ".join(k.split())


class SupabaseAdapter(DatabaseAdapter):
    def __init__(self, url: str, key: str):
        from supabase import create_client
        self.client = create_client(url, key)

    # ... existing methods ...

    def get_user_by_username(self, username: str) -> Optional[Dict[str, Any]]:
        resp = self.client.table("users").select("*").eq("username", username).limit(1).execute()
        data = resp.data or []
        return data[0] if data else None

    def get_user_by_id(self, user_id: str) -> Optional[Dict[str, Any]]:
        try:
            resp = self.client.table("users").select("*").eq("id", int(user_id)).limit(1).execute()
            data = resp.data or []
            return data[0] if data else None
        except Exception:
            return None

    def get_all_users(self) -> List[Dict[str, Any]]:
        resp = self.client.table("users").select("id, username, role, created_at").order("username").execute()
        return resp.data or []

    def add_user(self, username: str, password_hash: str, role: str = 'user') -> Dict[str, Any]:
        try:
            resp = self.client.table("users").insert({
                "username": username,
                "password_hash": password_hash,
                "role": role
            }).execute()
            return {"success": True, "id": resp.data[0]['id'] if resp.data else None}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def delete_user(self, user_id: int) -> Dict[str, Any]:
        try:
            self.client.table("users").delete().eq("id", user_id).execute()
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}
        resp = self.client.table("checklist_items").select("*").order("category").execute()
        data = resp.data or []
        return [
            {
                'category': r.get('category', ''),
                'item': r.get('item', ''),
                'key': r.get('key', ''),
                'is_checkbox': r.get('is_checkbox', True)
            }
            for r in data
        ]

    def get_all_dvr_risks(self) -> List[Dict[str, Any]]:
        resp = self.client.table("dvr_risks").select("*").order("ordine").execute()
        data = resp.data or []
        return [
            {
                'luogo': r.get('luogo', ''),
                'rischio': r.get('rischio', ''),
                'azione': r.get('azione', ''),
                'entro': r.get('entro', ''),
                'key': r.get('key', ''),
                'ordine': r.get('ordine', 999999),
                'image': r.get('image_url', '')
            }
            for r in data
        ]

    def match_risks(self, checked_keys: List[str]) -> List[Dict[str, Any]]:
        risks = self.get_all_dvr_risks()
        matched = []
        used = set()
        norm_ck = [self._normalize(k) for k in checked_keys]
        for i, ck in enumerate(checked_keys):
            n = norm_ck[i]
            if not n:
                continue
            for j, r in enumerate(risks):
                if j in used:
                    continue
                nr = self._normalize(r['key'])
                if nr == n or n in nr or nr in n:
                    matched.append({**r, 'key': ck})
                    used.add(j)
                    break
        matched.sort(key=lambda x: x.get('ordine', 999999))
        return matched

    def add_database_risk(self, risk: Dict[str, Any]) -> Dict[str, Any]:
        ord_val = risk.get('ordine', 999999)
        try:
            ordine = int(ord_val) if ord_val is not None else 999999
        except:
            ordine = 999999
        key_val = risk.get('rischio') or risk.get('luogo') or ""
        data = {
            "luogo": risk.get('luogo', ''),
            "rischio": risk.get('rischio', ''),
            "azione": risk.get('azione', ''),
            "entro": risk.get('entro', ''),
            "key": key_val,
            "ordine": ordine,
            "image_url": risk.get('image', '')
        }
        result = self.client.table("dvr_risks").insert(data).execute()
        return {"success": True, "message": "Criticità salvata su Supabase.", "id": result.data[0]['id'] if result.data else None}

    @staticmethod
    def _normalize(k):
        if not k:
            return ""
        k = str(k).lower().strip()
        k = re.sub(r'\(.*?\)', '', k)
        k = re.sub(r'[^a-z0-9\s]', ' ', k)
        return " ".join(k.split())

    def _load_users(self) -> Dict[str, Any]:
        if not os.path.exists(self._users_path):
            return {"users": []}
        try:
            with open(self._users_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {"users": []}

    def _save_users(self, data: Dict[str, Any]) -> None:
        os.makedirs(os.path.dirname(self._users_path), exist_ok=True)
        with open(self._users_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def get_user_by_username(self, username: str) -> Optional[Dict[str, Any]]:
        data = self._load_users()
        for u in data.get("users", []):
            if u.get("username") == username:
                return u
        return None

    def get_user_by_id(self, user_id: str) -> Optional[Dict[str, Any]]:
        try:
            uid = int(user_id)
            data = self._load_users()
            for u in data.get("users", []):
                if u.get("id") == uid:
                    return u
        except Exception:
            pass
        return None

    def get_all_users(self) -> List[Dict[str, Any]]:
        data = self._load_users()
        users = data.get("users", [])
        return [
            {
                "id": u.get("id"),
                "username": u.get("username", ""),
                "role": u.get("role", "user"),
                "created_at": u.get("created_at", "")
            }
            for u in users
        ]

    def add_user(self, username: str, password_hash: str, role: str = 'user') -> Dict[str, Any]:
        data = self._load_users()
        users = data.get("users", [])
        if any(u.get("username") == username for u in users):
            return {"success": False, "error": "Username già esistente."}
        new_id = (max((u.get("id", 0) for u in users), default=0) + 1)
        users.append({
            "id": new_id,
            "username": username,
            "password_hash": password_hash,
            "role": role,
            "created_at": __import__('datetime').datetime.now().isoformat()
        })
        data["users"] = users
        self._save_users(data)
        return {"success": True, "id": new_id}

    def delete_user(self, user_id: int) -> Dict[str, Any]:
        try:
            data = self._load_users()
            users = [u for u in data.get("users", []) if u.get("id") != user_id]
            data["users"] = users
            self._save_users(data)
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}


def get_adapter(backend: str = "local", **kwargs) -> DatabaseAdapter:
    if backend == "supabase":
        url = kwargs.get("supabase_url") or kwargs.get("url") or ""
        key = kwargs.get("supabase_key") or kwargs.get("key") or ""
        if not url or not key:
            raise ValueError("Supabase URL e KEY richiesti per backend='supabase'")
        return SupabaseAdapter(url, key)
    elif backend == "local":
        db_path = kwargs.get("db_path", "DATABASE.xlsx")
        return ExcelAdapter(db_path)
    else:
        raise ValueError(f"Backend '{backend}' non supportato. Usa 'local' o 'supabase'.")
