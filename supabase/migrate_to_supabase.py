"""
Script di migrazione dati da DATABASE.xlsx a Supabase.
Legge i due fogli Excel e popola le tabelle PostgreSQL su Supabase.
Uso:
    python supabase/migrate_to_supabase.py
env vars richieste:
    SUPABASE_URL, SUPABASE_KEY
"""
import os
import sys
import base64
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# Aggiungi la root del progetto al path per importare app.py
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl

load_dotenv()

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
DB_PATH = Path(__file__).resolve().parent.parent / "DATABASE.xlsx"


def extract_excel_images_as_base64(excel_path):
    """Estrae immagini da Excel nel formato {excel_row: base64_data_uri}."""
    images_map = {}
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            if 'xl/drawings/drawing9.xml' not in z.namelist():
                return images_map
            rels_xml = z.read('xl/drawings/_rels/drawing9.xml.rels').decode('utf-8')
            root_rels = ET.fromstring(rels_xml)
            rid_to_img_name = {}
            for rel in root_rels:
                r_id = rel.attrib['Id']
                target = rel.attrib['Target'].replace('../', 'xl/')
                rid_to_img_name[r_id] = target

            d9_xml = z.read('xl/drawings/drawing9.xml').decode('utf-8')
            root_d9 = ET.fromstring(d9_xml)
            ns = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            for anchor in root_d9.findall('xdr:oneCellAnchor', ns):
                from_elem = anchor.find('xdr:from', ns)
                blip_elem = anchor.find('.//a:blip', ns)
                if from_elem is not None and blip_elem is not None:
                    r_elem = from_elem.find('xdr:row', ns)
                    embed_id = blip_elem.attrib.get(
                        '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed'
                    )
                    if r_elem is not None and embed_id:
                        excel_row = int(r_elem.text) + 1
                        img_path = rid_to_img_name.get(embed_id)
                        if img_path and img_path in z.namelist():
                            img_bytes = z.read(img_path)
                            mime = 'image/png' if img_path.lower().endswith('.png') else 'image/jpeg'
                            b64 = base64.b64encode(img_bytes).decode('utf-8')
                            images_map[excel_row] = f"data:{mime};base64,{b64}"
    except Exception as e:
        print(f"Errore estrazione immagini: {e}")
    return images_map


def migrate_checklist(wb):
    """Migra il foglio CHECK LIST."""
    if 'CHECK LIST' not in wb.sheetnames:
        print("Foglio CHECK LIST non trovato, skip.")
        return []
    ws = wb['CHECK LIST']
    items = []
    seen = set()
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx < 4:
            continue
        categoria = row[0] or "GENERALE"
        nome_item = row[1]
        if nome_item is None:
            continue
        is_checkbox = isinstance(row[2], bool)
        key = row[6] if row[6] is not None else f"{categoria} {nome_item}"
        key = str(key).strip()
        if key in seen:
            continue
        seen.add(key)
        items.append({
            "category": str(categoria).strip(),
            "item": str(nome_item).strip(),
            "key": key,
            "is_checkbox": is_checkbox
        })
    return items


def migrate_dvr_risks(wb, images_map):
    """Migra il foglio DVR - RISCHI."""
    if 'DVR - RISCHI' not in wb.sheetnames:
        print("Foglio DVR - RISCHI non trovato, skip.")
        return []
    ws = wb['DVR - RISCHI']
    risks = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx < 2:
            continue
        excel_row_num = r_idx + 1
        luogo = row[0] or ""
        rischio = row[1] or ""
        azione = row[2] or ""
        entro = row[4] or ""
        key = row[5] or ""
        raw_ordine = row[7] if len(row) > 7 else None
        try:
            ordine = int(raw_ordine) if raw_ordine is not None else 999999
        except (ValueError, TypeError):
            ordine = 999999
        is_falso = row[6] if len(row) > 6 else "FALSO"
        is_falso_bool = str(is_falso).upper() == "FALSO"
        image_url = ""
        if excel_row_num in images_map:
            image_url = images_map[excel_row_num]
        risks.append({
            "luogo": str(luogo).strip(),
            "rischio": str(rischio).strip(),
            "azione": str(azione).strip(),
            "entro": str(entro).strip(),
            "key": str(key).strip(),
            "ordine": ordine,
            "image_url": image_url,
            "is_falso": is_falso_bool
        })
    return risks


def upload_image_to_storage(supabase: Client, data_uri: str, risk_id: int, row_num: int):
    """Carica immagine da data URI su Supabase Storage e restituisce il path."""
    if not data_uri.startswith("data:"):
        return ""
    try:
        header, b64_data = data_uri.split(",", 1)
        mime = header.split(":")[1].split(";")[0]
        ext = "png" if "png" in mime else "jpg"
        file_bytes = base64.b64decode(b64_data)
        path = f"risk-images/{risk_id}_{row_num}.{ext}"
        supabase.storage.from_("risk-images").upload(
            path,
            file_bytes,
            {"content-type": mime, "upsert": "true"}
        )
        public_url = supabase.storage.from_("risk-images").get_public_url(path)
        return public_url
    except Exception as e:
        print(f"Errore upload immagine riga {row_num}: {e}")
        return ""


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERRORE: Imposta SUPABASE_URL e SUPABASE_KEY nelle variabili d'ambiente.")
        sys.exit(1)
    if not DB_PATH.exists():
        print(f"ERRORE: DATABASE.xlsx non trovato in {DB_PATH}")
        sys.exit(1)

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(DB_PATH, data_only=True)

    print("Estrazione immagini da Excel...")
    images_map = extract_excel_images_as_base64(DB_PATH)
    print(f"Immagini estratte: {len(images_map)}")

    print("Migrazione CHECK LIST...")
    checklist_items = migrate_checklist(wb)
    print(f"Voci checklist trovate: {len(checklist_items)}")
    if checklist_items:
        # Svuota e reinserisce per evitare conflitti di ON CONFLICT
        supabase.table("checklist_items").delete().neq("id", 0).execute()
        batch_size = 500
        for i in range(0, len(checklist_items), batch_size):
            batch = checklist_items[i:i+batch_size]
            supabase.table("checklist_items").insert(batch).execute()
            print(f"  Checklist batch {i//batch_size + 1} inserito ({len(batch)} voci).")

    print("Migrazione DVR - RISCHI...")
    dvr_risks = migrate_dvr_risks(wb, images_map)
    print(f"Rischi DVR trovati: {len(dvr_risks)}")
    if dvr_risks:
        # Svuota e reinserisce per evitare conflitti
        supabase.table("dvr_risks").delete().neq("id", 0).execute()
        inserted_risks = []
        batch_size = 200
        for i in range(0, len(dvr_risks), batch_size):
            batch = dvr_risks[i:i+batch_size]
            result = supabase.table("dvr_risks").insert(batch).execute()
            inserted_risks.extend(result.data)
            print(f"  DVR batch {i//batch_size + 1} inserito ({len(batch)} rischi).")

        print("Caricamento immagini su Supabase Storage...")
        for i, risk in enumerate(inserted_risks):
            if risk.get("image_url"):
                original_row = dvr_risks[i]
                excel_row = original_row.get("id", i + 3)
                public_url = upload_image_to_storage(
                    supabase, risk["image_url"], risk["id"], excel_row
                )
                if public_url:
                    supabase.table("dvr_risks").update(
                        {"image_url": public_url}
                    ).eq("id", risk["id"]).execute()
                    print(f"  Immagine caricata per rischio id={risk['id']}")

    print("Migrazione completata!")


if __name__ == "__main__":
    main()
